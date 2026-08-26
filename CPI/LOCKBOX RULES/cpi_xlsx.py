"""Shared Excel formatting for CPI reports."""
from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Border, Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

REPORT_FONT = Font(name="Arial", size=10)


def apply_arial_10(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = REPORT_FONT


def apply_no_cell_borders(ws: Worksheet) -> None:
    none = Border()
    for row in ws.iter_rows():
        for cell in row:
            cell.border = none


def save_workbook(
    wb: Workbook,
    path,
    *,
    show_grid_lines: bool = True,
    cell_borders: bool = True,
) -> None:
    for ws in wb.worksheets:
        apply_arial_10(ws)
        if not cell_borders:
            apply_no_cell_borders(ws)
        ws.sheet_view.showGridLines = show_grid_lines
    wb.save(path)


def write_dict_rows_xlsx(
    path: Path,
    fieldnames: list[str],
    rows: list[dict],
    *,
    sheet_title: str = "Sheet1",
) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit(
            "openpyxl required for .xlsx output: pip install openpyxl"
        ) from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for c, col in enumerate(fieldnames, start=1):
        ws.cell(1, c, col)
    for r, row in enumerate(rows, start=2):
        for c, col in enumerate(fieldnames, start=1):
            val = row.get(col, "")
            if col == "Transaction ID" and val != "":
                try:
                    val = int(str(val).strip())
                except ValueError:
                    pass
            elif col in ("Check Amount",) and val != "":
                try:
                    val = float(str(val).replace(",", ""))
                except ValueError:
                    pass
            ws.cell(r, c, val if val != "" else None)
    path.parent.mkdir(parents=True, exist_ok=True)
    save_workbook(wb, path)
