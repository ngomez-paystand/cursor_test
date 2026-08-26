#!/usr/bin/env python3
"""
Build lockbox review Excel from Paystand invoice export + tif_review_queue.csv.

Includes every invoice export row (same order as the CSV), then the Paystand export
footer (row count + total amount) as the last spreadsheet row.

Good? rules:
  Invoice Number 0/empty (full scan — merchant + amount):
    Needs Human? = no  -> y
    Needs Human? = yes -> n
    Reason starts with "Not a check" (TIF confidently has no check image at all,
      e.g. a mailed notice/letter/legal filing/envelope) -> not a check
    No queue row for Transaction Id -> Good? left blank
  Invoice Number present (misroute-only check; amount not evaluated):
    Needs Human? = yes (OCR payee matches a different known merchant) -> n
    Anything else (matched, ambiguous, illegible, no queue row) -> Good? left blank
    (blank here means "not flagged", not "confirmed good" — amount/merchant were not
    fully re-verified since the invoice number is already on file)

Queue rows are matched by (Transaction Id, Invoice Number) so a check that has both an
invoice-0 line and an invoiced line doesn't mix up the two scan types.

Output columns match the manual lockbox spreadsheet (13 columns).
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from queue_tif_review import (
    discover_invoice_and_images,
    ensure_run_dir_exports_commas_clean,
    format_comma_audit_summary,
    is_missing_invoice,
    load_invoice_export_footer,
    load_invoice_export_rows,
    merchant_lookup_merged,
)
from cpi_xlsx import save_workbook

OUTPUT_COLUMNS = [
    "Mail Stop",
    "Merchant",
    "Invoice Number",
    "Good?",
    "Transaction Id",
    "Check Amount",
    "Deposit Date",
    "Routing Transit",
    "DDA",
    "Check Number",
    "Payer",
    "Merchant Account Number",
    "Apply Amount",
]


def _good_from_needs_human(needs_human: str, reason: str) -> str:
    if (reason or "").strip().lower().startswith("not a check"):
        return "not a check"
    t = (needs_human or "").strip().lower()
    if t == "no":
        return "y"
    if t == "yes":
        return "n"
    return ""


def load_queue_needs_human(queue_csv: Path) -> dict[tuple[str, str], tuple[str, str]]:
    """(Transaction ID, Invoice Number) -> (Needs Human?, Reason) (last row wins if duplicates).

    Keyed by the pair (not Transaction ID alone) because a single check can have both an
    invoice-0 line (full scan) and an invoiced line (misroute-only scan) in tif_review_queue.csv;
    keying by Transaction ID alone would let one scan type's result overwrite the other's.
    """
    out: dict[tuple[str, str], tuple[str, str]] = {}
    with queue_csv.open(newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        tid_key = None
        inv_key = None
        nh_key = None
        reason_key = None
        for name in reader.fieldnames or []:
            if name.replace(" ", "").lower() in ("transactionid",):
                tid_key = name
            if name.strip().lower() == "invoice number":
                inv_key = name
            if name.strip().lower() == "needs human?":
                nh_key = name
            if name.strip().lower() == "reason":
                reason_key = name
        if not tid_key or not nh_key:
            raise SystemExit(
                f"Queue CSV missing Transaction ID or Needs Human? columns: {queue_csv}"
            )
        for row in reader:
            tid = (row.get(tid_key) or "").strip()
            inv = (row.get(inv_key) or "").strip() if inv_key else ""
            if tid:
                reason = (row.get(reason_key) or "").strip() if reason_key else ""
                out[(tid, inv)] = ((row.get(nh_key) or "").strip(), reason)
    return out


def build_rows(
    invoice_csv: Path,
    queue_by_tid: dict[tuple[str, str], tuple[str, str]],
    merchant_lookup: dict[str, str],
) -> list[dict[str, str]]:
    rows_out: list[dict[str, str]] = []
    invoice_rows = load_invoice_export_rows(invoice_csv)
    if not invoice_rows:
        raise SystemExit(f"Empty invoice CSV: {invoice_csv}")
    for row in invoice_rows:
            ms = (row.get("Mail Stop") or "").strip()
            tid = (row.get("Transaction Id") or "").strip()
            inv_num = str(row.get("Invoice Number") or "").strip()
            merchant = merchant_lookup.get(ms, "")
            missing_inv = is_missing_invoice(inv_num)
            needs_human, reason = queue_by_tid.get((tid, inv_num), ("", ""))
            if missing_inv:
                good = _good_from_needs_human(needs_human, reason)
            else:
                # Misroute-only check: only a confirmed "yes" (different known merchant found)
                # is worth surfacing; anything else stays blank (not confirmed good).
                good = "n" if needs_human.strip().lower() == "yes" else ""
            rows_out.append(
                {
                    "Mail Stop": ms,
                    "Merchant": merchant,
                    "Invoice Number": str(row.get("Invoice Number") or "").strip(),
                    "Good?": good,
                    "Transaction Id": tid,
                    "Check Amount": (row.get("Check Amount") or "").strip(),
                    "Deposit Date": (row.get("Deposit Date") or "").strip(),
                    "Routing Transit": (row.get("Routing Transit") or "").strip(),
                    "DDA": (row.get("DDA") or "").strip(),
                    "Check Number": (row.get("Check Number") or "").strip(),
                    "Payer": (row.get("Payer") or "").strip(),
                    "Merchant Account Number": (
                        row.get("Merchant Account Number") or ""
                    ).strip(),
                    "Apply Amount": (row.get("Apply Amount") or "").strip(),
                }
            )
    return rows_out


def footer_row(count: str, total: str) -> dict[str, str]:
    """Lockbox footer mirroring Paystand invoice export last line (count, total)."""
    return {
        "Mail Stop": count,
        "Merchant": "",
        "Invoice Number": "",
        "Good?": "",
        "Transaction Id": "",
        "Check Amount": total,
        "Deposit Date": "",
        "Routing Transit": "",
        "DDA": "",
        "Check Number": "",
        "Payer": "",
        "Merchant Account Number": "",
        "Apply Amount": "",
        "_footer": "1",
    }


def _cell_value(col: str, val: str):
    if val == "":
        return None
    if col in ("Mail Stop", "Invoice Number"):
        try:
            return float(val) if "." in val else int(val)
        except ValueError:
            return val
    if col == "Transaction Id":
        try:
            return int(val)
        except ValueError:
            return val
    if col in ("Check Amount", "Apply Amount"):
        try:
            return float(val.replace(",", ""))
        except ValueError:
            return val
    if col == "Routing Transit":
        try:
            return int(val.replace(",", "").replace(" ", ""))
        except ValueError:
            return val
    return val


def write_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit(
            "openpyxl required for .xlsx output: pip install openpyxl"
        ) from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(1, c, col)
    for r, row in enumerate(rows, start=2):
        for c, col in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(r, c, _cell_value(col, (row.get(col) or "").strip()))
    path.parent.mkdir(parents=True, exist_ok=True)
    save_workbook(wb, path, show_grid_lines=False, cell_borders=False)


def main() -> None:
    here = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(
        description="Build lockbox Good? Excel (all invoice rows; Good? only for invoice 0)."
    )
    ap.add_argument(
        "--run-dir",
        type=Path,
        required=True,
        help="Day folder (e.g. JUNE/06-04-2026) with invoice CSV and tif_review_queue.csv",
    )
    ap.add_argument(
        "--invoice-csv",
        type=Path,
        default=None,
        help="Override invoice export path",
    )
    ap.add_argument(
        "--queue-csv",
        type=Path,
        default=None,
        help="Override tif_review_queue.csv (default: <run-dir>/tif_review_queue.csv)",
    )
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx (default: <run-dir>/lockbox_report.xlsx)",
    )
    args = ap.parse_args()

    run_dir = args.run_dir.resolve()
    # Comma audit never rewrites the raw exports (auto_fix=False, always — see
    # cpi-lockbox-comma-audit.mdc). ANY issue found stops processing immediately.
    comma_reports = ensure_run_dir_exports_commas_clean(
        run_dir,
        auto_fix=False,
    )
    any_issue = any(r.misaligned for r in comma_reports)
    if any_issue:
        print(format_comma_audit_summary(comma_reports, run_dir=run_dir), flush=True)
        print(flush=True)
        raise SystemExit(
            "Aborting: comma audit found issue(s) in the raw export(s) above. Files are "
            "left untouched — fix the raw invoice/check/image-metadata CSV(s) by hand and "
            "re-run before generating the lockbox report."
        )
    if args.invoice_csv is None:
        invoice_csv, _ = discover_invoice_and_images(run_dir)
    else:
        invoice_csv = args.invoice_csv.resolve()
    queue_csv = (args.queue_csv or run_dir / "tif_review_queue.csv").resolve()
    if not queue_csv.is_file():
        raise SystemExit(f"Queue CSV not found: {queue_csv}")
    out_path = (args.output or run_dir / "lockbox_report.xlsx").resolve()

    lookup_dirs = [run_dir, invoice_csv.parent, here]
    merchant_lookup = merchant_lookup_merged(lookup_dirs)
    queue_by_tid = load_queue_needs_human(queue_csv)
    rows = build_rows(invoice_csv, queue_by_tid, merchant_lookup)
    footer = load_invoice_export_footer(invoice_csv)
    if footer:
        rows.append(footer_row(*footer))
    write_xlsx(rows, out_path)

    data_rows = [r for r in rows if not r.get("_footer")]
    inv0_rows = [r for r in data_rows if is_missing_invoice(r.get("Invoice Number", ""))]
    y_count = sum(1 for r in inv0_rows if r.get("Good?") == "y")
    n_count = sum(1 for r in inv0_rows if r.get("Good?") == "n")
    not_check_count = sum(1 for r in inv0_rows if r.get("Good?") == "not a check")
    inv0 = len(inv0_rows)
    misroute_n = sum(
        1
        for r in data_rows
        if not is_missing_invoice(r.get("Invoice Number", "")) and r.get("Good?") == "n"
    )
    footer_note = ""
    if footer:
        footer_note = f", footer={footer[0]}/{footer[1]}"
    print(
        f"Wrote: {out_path} ({len(data_rows)} rows{footer_note}, invoice 0={inv0}, "
        f"Good?: y={y_count}, n={n_count}, not a check={not_check_count}, "
        f"invoiced misroute flagged={misroute_n}, other rows Good? blank)"
    )


if __name__ == "__main__":
    main()
